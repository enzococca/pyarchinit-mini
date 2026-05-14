"""Pottery web routes. Call _register_pottery_routes(app) from create_app()."""
from __future__ import annotations

from flask import (
    render_template, request, redirect, url_for, flash, jsonify, abort,
)
from flask_login import login_required

from ..services.pottery_service import PotteryService
from ..services.pottery_dto import PotteryDTO


_POTTERY_FORM_FIELDS = (
    "id_number", "sito", "area", "us", "box", "photo", "drawing", "anno",
    "fabric", "percent", "material", "form", "specific_form", "ware",
    "munsell", "surf_trat", "exdeco", "intdeco", "wheel_made",
    "descrip_ex_deco", "descrip_in_deco", "note", "diametro_max", "qty",
    "diametro_rim", "diametro_bottom", "diametro_height",
    "diametro_preserved", "specific_shape", "bag", "sector",
)
_INT_FIELDS = {"id_number", "us", "box", "anno", "qty", "bag"}
_NUM_FIELDS = {
    "diametro_max", "diametro_rim", "diametro_bottom",
    "diametro_height", "diametro_preserved",
}


def _extract_pottery_form(form) -> dict:
    """Pull pottery fields from a Flask request.form, coercing types."""
    out = {}
    for k in _POTTERY_FORM_FIELDS:
        v = form.get(k)
        if v in (None, ""):
            continue
        if k in _INT_FIELDS:
            try:
                out[k] = int(v)
            except ValueError:
                continue
        elif k in _NUM_FIELDS:
            try:
                out[k] = float(v)
            except ValueError:
                continue
        else:
            out[k] = v
    return out


def _distinct_values(app, column_name: str):
    """Return distinct values for a Pottery column (DB + THESAURUS_MAPPINGS) as JSON."""
    from ..models.pottery import Pottery
    from ..models.thesaurus import THESAURUS_MAPPINGS
    with app.db_manager.connection.get_session() as session:
        col = getattr(Pottery, column_name)
        rows = session.query(col).filter(col.isnot(None)).distinct().all()
        db_values = {r[0] for r in rows if r[0]}
    defaults = set(THESAURUS_MAPPINGS.get("pottery_table", {}).get(column_name, []))
    merged = sorted(db_values | defaults)
    return jsonify({"values": merged})


def _register_pottery_routes(app):
    """Register pottery URL rules on the Flask app."""

    @app.route("/pottery")
    @login_required
    def pottery_list():
        page = int(request.args.get("page", 1))
        size = int(request.args.get("size", 25))
        filters = {
            k: request.args.get(k)
            for k in ("sito", "area", "us", "form", "fabric", "ware", "q")
            if request.args.get(k)
        }
        svc = PotteryService(app.db_manager)
        items, total = svc.get_all_pottery(page=page, size=size, filters=filters)
        ids = [p.id_rep for p in items]
        media_ids: set[int] = set()
        if ids:
            from ..models.media import Media
            with app.db_manager.connection.get_session() as session:
                rows = (
                    session.query(Media.entity_id)
                    .filter(Media.entity_type == "pottery", Media.entity_id.in_(ids))
                    .distinct()
                    .all()
                )
                media_ids = {r[0] for r in rows}
        return render_template(
            "pottery/list.html",
            items=[PotteryDTO.from_model(p) for p in items],
            total=total, page=page, size=size, filters=filters,
            media_ids=media_ids,
        )

    @app.route("/pottery/create", methods=["GET", "POST"])
    @login_required
    def pottery_create():
        svc = PotteryService(app.db_manager)
        if request.method == "POST":
            data = _extract_pottery_form(request.form)
            try:
                p = svc.create_pottery(data)
            except ValueError as e:
                flash(str(e), "danger")
                return render_template(
                    "pottery/form.html",
                    pottery=None, form_data=data, mode="create",
                )
            flash(f"Pottery #{p.id_rep} created.", "success")
            try:
                return redirect(url_for("pottery_detail", id_rep=p.id_rep))
            except Exception:
                # pottery_detail not yet registered (during T11 development) — go to list
                return redirect(url_for("pottery_list"))
        return render_template(
            "pottery/form.html", pottery=None, form_data={}, mode="create"
        )

    @app.route("/pottery/<int:id_rep>")
    @login_required
    def pottery_detail(id_rep: int):
        svc = PotteryService(app.db_manager)
        p = svc.get_pottery_by_id(id_rep)
        if not p:
            abort(404)
        return render_template(
            "pottery/detail.html",
            pottery=PotteryDTO.from_model(p),
        )

    @app.route("/pottery/<int:id_rep>/edit", methods=["GET", "POST"])
    @login_required
    def pottery_edit(id_rep: int):
        svc = PotteryService(app.db_manager)
        p = svc.get_pottery_by_id(id_rep)
        if not p:
            abort(404)

        def _media_for_pottery():
            try:
                from ..services.media_service import MediaService
                msvc = MediaService(app.db_manager)
                return msvc.get_media_by_entity("pottery", id_rep)
            except Exception:
                return []

        if request.method == "POST":
            data = _extract_pottery_form(request.form)
            try:
                svc.update_pottery(id_rep, data)
            except ValueError as e:
                flash(str(e), "danger")
                return render_template(
                    "pottery/form.html",
                    pottery=PotteryDTO.from_model(p),
                    form_data=data, mode="edit",
                    media_list=_media_for_pottery(),
                )
            flash(f"Pottery #{id_rep} updated.", "success")
            return redirect(url_for("pottery_detail", id_rep=id_rep))
        return render_template(
            "pottery/form.html",
            pottery=PotteryDTO.from_model(p),
            form_data={}, mode="edit",
            media_list=_media_for_pottery(),
        )

    @app.route("/pottery/<int:id_rep>/delete", methods=["POST"])
    @login_required
    def pottery_delete(id_rep: int):
        svc = PotteryService(app.db_manager)
        if not svc.delete_pottery(id_rep):
            abort(404)
        flash(f"Pottery #{id_rep} deleted.", "info")
        return redirect(url_for("pottery_list"))

    @app.route("/api/pottery/forms")
    @login_required
    def pottery_api_forms():
        return _distinct_values(app, "form")

    @app.route("/api/pottery/fabrics")
    @login_required
    def pottery_api_fabrics():
        return _distinct_values(app, "fabric")

    @app.route("/api/pottery/wares")
    @login_required
    def pottery_api_wares():
        return _distinct_values(app, "ware")

    @app.route("/api/media/pottery")
    @login_required
    def api_media_pottery():
        """Get list of all pottery items for media upload form."""
        svc = PotteryService(app.db_manager)
        items, _ = svc.get_all_pottery(page=1, size=500)
        return jsonify({
            "items": [
                {
                    "id_rep": p.id_rep,
                    "sito": p.sito,
                    "id_number": p.id_number,
                    "form": p.form,
                    "fabric": p.fabric,
                }
                for p in items
            ]
        })

    @app.route("/pottery/<int:id_rep>/media")
    @login_required
    def pottery_media(id_rep: int):
        """Redirect to the universal media manager for this pottery record."""
        try:
            return redirect(url_for("media_manage", entity_type="pottery", entity_id=id_rep))
        except Exception:
            try:
                return redirect(url_for("upload_media", entity_type="pottery", entity_id=id_rep))
            except Exception:
                abort(404)

    @app.route("/export/pottery/excel")
    @login_required
    def pottery_export_excel():
        import io as _io
        import pandas as pd
        from datetime import datetime, timezone
        from flask import send_file
        svc = PotteryService(app.db_manager)
        filters = {k: request.args.get(k) for k in ("sito","area","us","form","fabric") if request.args.get(k)}
        items, _ = svc.get_all_pottery(page=1, size=10_000, filters=filters)
        rows = [PotteryDTO.from_model(p).to_dict() for p in items]
        df = pd.DataFrame(rows, columns=list(_POTTERY_FORM_FIELDS) + ["id_rep"])
        meta = pd.DataFrame([{
            "exported_at": datetime.now(timezone.utc).isoformat(),
            "version": "2.1.60", "filters": str(filters),
            "row_count": len(rows),
        }])
        buf = _io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            df.to_excel(w, sheet_name="pottery", index=False)
            meta.to_excel(w, sheet_name="metadata", index=False)
        buf.seek(0)
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name="pottery.xlsx",
        )

    @app.route("/export/pottery/csv")
    @login_required
    def pottery_export_csv():
        import io as _io
        import csv
        from flask import Response
        svc = PotteryService(app.db_manager)
        filters = {k: request.args.get(k) for k in ("sito","area","us","form","fabric") if request.args.get(k)}
        items, _ = svc.get_all_pottery(page=1, size=10_000, filters=filters)
        header = list(_POTTERY_FORM_FIELDS) + ["id_rep"]
        buf = _io.StringIO()
        w = csv.DictWriter(buf, fieldnames=header)
        w.writeheader()
        for p in items:
            d = PotteryDTO.from_model(p).to_dict()
            w.writerow({k: d.get(k, "") for k in header})
        return Response(
            buf.getvalue(), mimetype="text/csv",
            headers={"Content-Disposition": 'attachment; filename="pottery.csv"'},
        )

    @app.route("/import/pottery", methods=["GET"])
    @login_required
    def pottery_import_form():
        return render_template("pottery/import_form.html")

    @app.route("/import/pottery/excel", methods=["POST"])
    @login_required
    def pottery_import_excel():
        import csv as _csv
        from openpyxl import load_workbook
        from ..models.pottery import Pottery as _Pottery
        f = request.files.get("file")
        if not f:
            flash("Missing file", "danger")
            return redirect(url_for("pottery_import_form"))
        mode = request.form.get("mode", "skip")  # skip | update | renumber

        # Parse file into list of dicts
        try:
            fname = (f.filename or "").lower()
            if fname.endswith(".csv"):
                text = f.read().decode("utf-8-sig")
                rows_raw = list(_csv.DictReader(text.splitlines()))
            else:
                wb = load_workbook(f, read_only=True, data_only=True)
                ws = wb["pottery"]
                row_iter = ws.iter_rows(values_only=True)
                headers = [str(h) if h is not None else "" for h in next(row_iter)]
                rows_raw = [dict(zip(headers, r)) for r in row_iter]
                wb.close()
        except Exception as e:
            flash(f"Cannot read file: {e}", "danger")
            return redirect(url_for("pottery_import_form"))

        def _coerce(k, v):
            """Coerce a raw cell value to the appropriate Python type."""
            if v is None or v == "":
                return None
            if k in _INT_FIELDS:
                try:
                    return int(float(v))
                except (TypeError, ValueError):
                    return None
            if k in _NUM_FIELDS:
                try:
                    return float(v)
                except (TypeError, ValueError):
                    return None
            return str(v) if not isinstance(v, str) else v

        svc = PotteryService(app.db_manager)
        stats = {"inserted": 0, "updated": 0, "skipped": 0, "errors": []}
        for idx, row in enumerate(rows_raw):
            data = {k: _coerce(k, v) for k, v in row.items() if k in _POTTERY_FORM_FIELDS}
            # Drop None values so service defaults apply
            data = {k: v for k, v in data.items() if v is not None}
            sito = data.get("sito")
            idn = data.get("id_number")
            if not sito:
                stats["errors"].append(f"row {idx+2}: missing sito")
                continue
            existing_id = None
            if idn is not None:
                with app.db_manager.connection.get_session() as session:
                    existing = (
                        session.query(_Pottery)
                        .filter(_Pottery.sito == sito, _Pottery.id_number == int(idn))
                        .first()
                    )
                    existing_id = existing.id_rep if existing is not None else None
            try:
                if existing_id is not None and mode == "skip":
                    stats["skipped"] += 1
                elif existing_id is not None and mode == "update":
                    svc.update_pottery(existing_id, data)
                    stats["updated"] += 1
                elif existing_id is not None and mode == "renumber":
                    data.pop("id_number", None)
                    svc.create_pottery(data)
                    stats["inserted"] += 1
                else:
                    svc.create_pottery(data)
                    stats["inserted"] += 1
            except ValueError as e:
                stats["errors"].append(f"row {idx+2}: {e}")

        flash(
            f"Import done: {stats['inserted']} inserted, "
            f"{stats['updated']} updated, {stats['skipped']} skipped, "
            f"{len(stats['errors'])} errors.",
            "info",
        )
        return redirect(url_for("pottery_list"))

    @app.route("/export/pottery_pdf")
    @login_required
    def pottery_export_pdf():
        from io import BytesIO
        from flask import send_file
        from ..services.pottery_pdf_service import PotteryPDFService
        from .. import __version__
        svc = PotteryService(app.db_manager)
        filters = {k: request.args.get(k) for k in ("sito","area","us","form","fabric") if request.args.get(k)}
        items, _ = svc.get_all_pottery(page=1, size=10_000, filters=filters)
        if not items:
            flash("No pottery records matching filters.", "warning")
            return redirect(url_for("pottery_list"))
        pdf_bytes = PotteryPDFService.render_sheets(items, version=__version__)
        return send_file(
            BytesIO(pdf_bytes), mimetype="application/pdf",
            as_attachment=True, download_name="pottery.pdf",
        )

    @app.route("/export/pottery_single_pdf/<int:id_rep>")
    @login_required
    def pottery_export_single_pdf(id_rep: int):
        from io import BytesIO
        from flask import send_file
        from ..services.pottery_pdf_service import PotteryPDFService
        from .. import __version__
        svc = PotteryService(app.db_manager)
        p = svc.get_pottery_by_id(id_rep)
        if not p:
            abort(404)
        pdf_bytes = PotteryPDFService.render_sheets([p], version=__version__)
        return send_file(
            BytesIO(pdf_bytes), mimetype="application/pdf",
            as_attachment=True, download_name=f"pottery_{id_rep}.pdf",
        )

    @app.route("/api/pottery/stats")
    @login_required
    def pottery_api_stats():
        from sqlalchemy import func as f
        from ..models.pottery import Pottery as _Pottery
        svc = PotteryService(app.db_manager)
        sito = request.args.get("sito")
        area = request.args.get("area")
        us = request.args.get("us")
        us_int = int(us) if us and us.isdigit() else None

        by_form = svc.get_form_distribution(sito) if sito else svc.get_form_distribution()
        by_fabric = svc.get_fabric_distribution(sito) if sito else svc.get_fabric_distribution()
        by_ware = svc.get_ware_distribution(sito) if sito else svc.get_ware_distribution()

        filters = {k: v for k, v in {"sito": sito, "area": area, "us": us_int}.items() if v}
        total = svc.count_pottery(filters or None)
        mni = svc.calculate_mni(sito, area, us_int)["total"] if sito else 0

        with app.db_manager.connection.get_session() as session:
            q = session.query(_Pottery.anno, f.count(_Pottery.id_rep)).group_by(_Pottery.anno)
            if sito:
                q = q.filter(_Pottery.sito == sito)
            by_anno = [{"anno": a, "count": c} for a, c in q.all() if a is not None]

        return jsonify({
            "total": total,
            "by_form": [{"form": k, "count": v} for k, v in by_form.items()],
            "by_fabric": [{"fabric": k, "count": v} for k, v in by_fabric.items()],
            "by_ware": [{"ware": k, "count": v} for k, v in by_ware.items()],
            "by_anno": by_anno,
            "mni": mni,
        })
