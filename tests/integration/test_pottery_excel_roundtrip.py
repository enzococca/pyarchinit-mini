"""Excel and CSV export/import round-trip tests for pottery."""
import io
from openpyxl import load_workbook

# Re-use the flask_app + client + pottery_service fixtures from the other test file
from tests.integration.test_pottery_routes import flask_app, client, pottery_service


def test_excel_export_returns_xlsx(client, pottery_service):
    pottery_service.create_pottery({"sito": "X", "form": "Olla", "fabric": "Coarse", "qty": 1})
    r = client.get("/export/pottery/excel")
    assert r.status_code == 200
    assert r.headers["Content-Type"].startswith("application/")
    wb = load_workbook(io.BytesIO(r.data))
    assert "pottery" in wb.sheetnames


def test_csv_export(client, pottery_service):
    pottery_service.create_pottery({"sito": "X", "form": "Olla"})
    r = client.get("/export/pottery/csv")
    assert r.status_code == 200
    body = r.data.decode("utf-8")
    assert "sito" in body and "Olla" in body


import pandas as pd


def test_excel_import_roundtrip(client, pottery_service):
    df = pd.DataFrame([{
        "sito": "X", "form": "Olla", "fabric": "Coarse", "qty": 1, "id_number": 10,
    }, {
        "sito": "X", "form": "Ciotola", "fabric": "Fine", "qty": 2, "id_number": 11,
    }])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="pottery", index=False)
    buf.seek(0)
    r = client.post(
        "/import/pottery/excel",
        data={"file": (buf, "in.xlsx"), "mode": "skip"},
        content_type="multipart/form-data",
    )
    assert r.status_code in (200, 302)
    items, total = pottery_service.get_all_pottery()
    assert total == 2


def test_excel_import_skip_duplicates(client, pottery_service):
    pottery_service.create_pottery({"sito": "X", "id_number": 10, "form": "Old"})
    df = pd.DataFrame([{"sito": "X", "id_number": 10, "form": "New"}])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="pottery", index=False)
    buf.seek(0)
    client.post(
        "/import/pottery/excel",
        data={"file": (buf, "in.xlsx"), "mode": "skip"},
        content_type="multipart/form-data",
    )
    items, _ = pottery_service.get_all_pottery()
    forms = [i.form for i in items]
    assert "Old" in forms
    assert "New" not in forms  # skip mode: existing not overwritten


def test_excel_import_update_mode(client, pottery_service):
    pottery_service.create_pottery({"sito": "X", "id_number": 10, "form": "Old"})
    df = pd.DataFrame([{"sito": "X", "id_number": 10, "form": "New"}])
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, sheet_name="pottery", index=False)
    buf.seek(0)
    client.post(
        "/import/pottery/excel",
        data={"file": (buf, "in.xlsx"), "mode": "update"},
        content_type="multipart/form-data",
    )
    items, _ = pottery_service.get_all_pottery()
    forms = [i.form for i in items]
    assert "New" in forms
